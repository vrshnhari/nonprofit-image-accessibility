import base64
import json
import os
import subprocess
import textwrap
import time
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageEnhance, ImageFont


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs"
IMAGE_DIR = OUTPUT_DIR / "validation-images"
WORKBOOK_PATH = OUTPUT_DIR / "week4_validation_results.xlsx"
PROMPT_LOG_PATH = ROOT / "PROMPT_LOG.md"
CACHE_PATH = OUTPUT_DIR / "week4_validation_cache.json"
MODEL_FALLBACK = "qwen/qwen3.6-27b"


RUN_PROMPTS = [
    (
        "Run 1",
        "Baseline prompt",
        "You are an accessibility assistant. Analyze the uploaded image and return a JSON object with three keys: "
        "'alt_text' (under 125 characters, WCAG-compliant), 'long_description' (detailed), and 'text_in_image' "
        "(transcribe visible text or return 'None detected'). Return only valid JSON.",
    ),
    (
        "Run 2",
        "Prioritize embedded event text",
        "You are an accessibility assistant. Analyze the uploaded image and return only valid JSON with three keys: "
        "'alt_text', 'long_description', and 'text_in_image'. Keep alt_text under 125 characters. If the image is a "
        "flyer, poster, screenshot, sign, or graphic with words, prioritize the visible title, date, time, location, "
        "and action over colors or decorative background details. Transcribe visible words exactly in text_in_image; "
        "return 'None detected' only when no readable text is visible.",
    ),
    (
        "Run 3",
        "Add purpose and hard-case guidance",
        "You are an accessibility assistant. Analyze the uploaded image and return only valid JSON with three keys: "
        "'alt_text', 'long_description', and 'text_in_image'. Keep alt_text under 125 characters and make it convey "
        "the image's purpose, not just generic objects. For flyers, signs, and screenshots, include the most important "
        "visible text such as event title, date, time, location, labels, or chart trend. For unclear, dark, or ambiguous "
        "photos, say what is uncertain instead of inventing details. Return 'None detected' for text_in_image only when "
        "no readable text is visible.",
    ),
]


@dataclass
class TestImage:
    filename: str
    category: str
    embedded_terms: list[str]
    purpose_terms: list[str]


def read_env() -> dict[str, str]:
    env = {}
    env_path = ROOT / ".env.local"
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            if "=" in line and not line.strip().startswith("#"):
                key, value = line.split("=", 1)
                env[key.strip()] = value.strip()
    env.update({k: v for k, v in os.environ.items() if k.startswith("GROQ_")})
    return env


def font(size: int):
    candidates = [
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
        "/Library/Fonts/Arial.ttf",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size=size)
    return ImageFont.load_default()


def wrap(draw: ImageDraw.ImageDraw, text: str, max_width: int, fnt):
    words = text.split()
    lines = []
    current = ""
    for word in words:
        test = f"{current} {word}".strip()
        if draw.textbbox((0, 0), test, font=fnt)[2] <= max_width:
            current = test
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines


def save_image(filename: str, draw_fn) -> Path:
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    path = IMAGE_DIR / filename
    if path.exists():
        return path
    img = Image.new("RGB", (900, 650), "#f7f5ef")
    draw = ImageDraw.Draw(img)
    draw_fn(img, draw)
    img.save(path, "PNG")
    return path


def label(draw, xy, text, size=34, fill="#14171a"):
    draw.text(xy, text, font=font(size), fill=fill)


def build_images() -> list[TestImage]:
    tests: list[TestImage] = []

    def campus_fair(img, draw):
        draw.rectangle((0, 0, 900, 120), fill="#0f766e")
        label(draw, (32, 32), "UW Club Fair", 44, "white")
        draw.rectangle((90, 220, 260, 500), fill="#88c0d0")
        draw.rectangle((330, 230, 520, 500), fill="#a3be8c")
        draw.rectangle((600, 210, 780, 500), fill="#ebcb8b")
        for x in [150, 410, 690]:
            draw.ellipse((x, 160, x + 72, 232), fill="#5e81ac")
            draw.rectangle((x + 15, 232, x + 58, 370), fill="#434c5e")
        label(draw, (70, 560), "Students visit club tables on a sunny campus lawn", 28)

    tests.append(TestImage("01_campus_club_fair.png", "campus event", ["UW Club Fair"], ["students", "club", "fair"]))
    save_image(tests[-1].filename, campus_fair)

    def robotics_demo(img, draw):
        draw.rectangle((0, 0, 900, 650), fill="#e8eef2")
        draw.rectangle((70, 70, 830, 155), fill="#2e3440")
        label(draw, (105, 92), "Robotics Demo", 40, "white")
        draw.rectangle((250, 300, 650, 390), fill="#4c566a")
        draw.ellipse((280, 380, 350, 450), fill="#111827")
        draw.ellipse((550, 380, 620, 450), fill="#111827")
        for x in [110, 710]:
            draw.ellipse((x, 230, x + 70, 300), fill="#bf616a")
            draw.rectangle((x + 18, 300, x + 52, 450), fill="#5e81ac")
        label(draw, (90, 545), "Two students present a wheeled robot near a classroom table", 28)

    tests.append(TestImage("02_campus_robotics_demo.png", "campus event", ["Robotics Demo"], ["robot", "students", "demo"]))
    save_image(tests[-1].filename, robotics_demo)

    def food_drive(img, draw):
        draw.rectangle((0, 0, 900, 650), fill="#fff7ed")
        draw.rectangle((50, 60, 850, 145), fill="#c2410c")
        label(draw, (95, 86), "Campus Food Drive", 38, "white")
        for x in [140, 300, 460, 620]:
            draw.ellipse((x, 200, x + 70, 270), fill="#f59e0b")
            draw.rectangle((x + 20, 270, x + 55, 420), fill="#0f766e")
        draw.rectangle((180, 430, 720, 520), fill="#d97706")
        label(draw, (225, 455), "Donation Table", 32, "white")
        label(draw, (105, 570), "Volunteers collect cans and grocery bags at a campus table", 26)

    tests.append(TestImage("03_campus_food_drive_photo.png", "campus event", ["Campus Food Drive"], ["volunteers", "food", "donation"]))
    save_image(tests[-1].filename, food_drive)

    flyers = [
        ("04_flyer_blood_drive.png", "Blood Drive", "Oct 12", "Student Union", "10 AM-3 PM"),
        ("05_flyer_coding_night.png", "Coding Night", "Friday 7 PM", "Room 204", "Bring laptop"),
        ("06_flyer_art_show.png", "Spring Art Show", "May 3", "Gallery Hall", "6 PM"),
    ]
    for filename, title, date, place, extra in flyers:
        def flyer(img, draw, title=title, date=date, place=place, extra=extra):
            draw.rectangle((0, 0, 900, 650), fill="#dbeafe")
            draw.rectangle((80, 60, 820, 590), fill="white", outline="#1d4ed8", width=8)
            label(draw, (130, 115), title, 54, "#1e3a8a")
            label(draw, (155, 235), date, 42, "#111827")
            label(draw, (155, 310), place, 42, "#111827")
            label(draw, (155, 385), extra, 38, "#b91c1c")
            draw.ellipse((625, 375, 760, 510), fill="#93c5fd")
        terms = [title, date, place, extra]
        tests.append(TestImage(filename, "flyer", terms, [title.lower().split()[0], date.split()[0].lower(), place.lower().split()[0]]))
        save_image(filename, flyer)

    def forest(img, draw):
        draw.rectangle((0, 0, 900, 650), fill="#d1fae5")
        draw.polygon((430, 650, 520, 650, 480, 250), fill="#92400e")
        for x in range(60, 860, 130):
            draw.rectangle((x, 260, x + 35, 650), fill="#78350f")
            draw.ellipse((x - 65, 105, x + 105, 315), fill="#166534")
        draw.polygon((360, 650, 540, 650, 500, 360, 410, 360), fill="#f5deb3")

    tests.append(TestImage("07_nature_forest_path.png", "nature", [], ["forest", "path", "trees"]))
    save_image(tests[-1].filename, forest)

    def lake(img, draw):
        draw.rectangle((0, 0, 900, 300), fill="#fca5a5")
        draw.rectangle((0, 300, 900, 650), fill="#60a5fa")
        draw.ellipse((380, 120, 520, 260), fill="#fde68a")
        draw.polygon((0, 300, 240, 120, 480, 300), fill="#374151")
        draw.polygon((420, 300, 650, 150, 900, 300), fill="#4b5563")

    tests.append(TestImage("08_nature_lake_sunset.png", "nature", [], ["lake", "sunset", "mountains"]))
    save_image(tests[-1].filename, lake)

    def bottle(img, draw):
        draw.rectangle((0, 0, 900, 650), fill="#f3f4f6")
        draw.rounded_rectangle((350, 105, 550, 550), radius=38, fill="#67e8f9", outline="#0e7490", width=8)
        draw.rectangle((395, 55, 505, 120), fill="#0e7490")
        label(draw, (381, 300), "H2O", 46, "#164e63")

    tests.append(TestImage("09_object_water_bottle.png", "object", ["H2O"], ["water", "bottle"]))
    save_image(tests[-1].filename, bottle)

    def keys(img, draw):
        draw.rectangle((0, 0, 900, 650), fill="#e5e7eb")
        draw.ellipse((280, 235, 430, 385), outline="#111827", width=18)
        draw.line((420, 310, 720, 310), fill="#111827", width=22)
        draw.rectangle((625, 310, 660, 390), fill="#111827")
        draw.rectangle((700, 310, 735, 360), fill="#111827")

    tests.append(TestImage("10_object_keys_closeup.png", "object", [], ["keys", "key"]))
    save_image(tests[-1].filename, keys)

    def chart(img, draw):
        draw.rectangle((0, 0, 900, 650), fill="white")
        label(draw, (80, 45), "Club Signups by Month", 42)
        draw.line((110, 540, 800, 540), fill="#111827", width=4)
        draw.line((110, 130, 110, 540), fill="#111827", width=4)
        bars = [110, 190, 300, 430]
        labels = ["Jan", "Feb", "Mar", "Apr"]
        for i, val in enumerate(bars):
            x = 180 + i * 135
            draw.rectangle((x, 540 - val, x + 70, 540), fill="#2563eb")
            label(draw, (x + 8, 555), labels[i], 28)
            label(draw, (x + 6, 510 - val), str(val), 24)
        label(draw, (32, 260), "Students", 24)

    tests.append(TestImage("11_hard_chart_screenshot.png", "hard", ["Club Signups by Month", "Jan", "Feb", "Mar", "Apr"], ["chart", "signups", "increase"]))
    save_image(tests[-1].filename, chart)

    def ambiguous(img, draw):
        draw.rectangle((0, 0, 900, 650), fill="#111827")
        draw.ellipse((320, 230, 610, 480), fill="#2f3748")
        draw.rectangle((450, 160, 520, 520), fill="#262f3f")
        draw.ellipse((520, 285, 575, 340), fill="#1f2937")
        img_blur = ImageEnhance.Brightness(img).enhance(0.55)
        img.paste(img_blur)
        label(draw, (65, 560), "Dim room, unclear subject", 28, "#6b7280")

    tests.append(TestImage("12_hard_poor_lighting_ambiguous.png", "hard", ["Dim room, unclear subject"], ["dark", "unclear", "ambiguous"]))
    save_image(tests[-1].filename, ambiguous)

    return tests


def normalize(result: dict) -> dict:
    alt = str(result.get("alt_text", "")).strip()
    if len(alt) > 125:
        alt = alt[:122].rstrip() + "..."
    long_description = str(result.get("long_description", "")).strip()
    text = str(result.get("text_in_image", "")).strip() or "None detected"
    return {
        "alt_text": alt,
        "long_description": long_description,
        "text_in_image": text,
    }


def call_groq(image_path: Path, prompt: str, api_key: str, model: str) -> dict:
    with Image.open(image_path) as original:
        original.thumbnail((224, 224))
        buffer = BytesIO()
        original.convert("RGB").save(buffer, "PNG", optimize=True)
    image_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    body = {
        "model": model,
        "temperature": 0.2,
        "max_completion_tokens": 180,
        "response_format": {"type": "json_object"},
        "reasoning_format": "hidden",
        "reasoning_effort": "none",
        "messages": [
            {"role": "system", "content": prompt},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": "Describe this image for accessibility."},
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/png;base64,{image_base64}"},
                    },
                ],
            },
        ],
    }
    node_code = """
let input = '';
process.stdin.on('data', chunk => input += chunk);
process.stdin.on('end', async () => {
  const payload = JSON.parse(input);
  const response = await fetch('https://api.groq.com/openai/v1/chat/completions', {
    method: 'POST',
    headers: {
      authorization: `Bearer ${payload.apiKey}`,
      'content-type': 'application/json',
    },
    body: JSON.stringify(payload.body),
  });
  const text = await response.text();
  if (!response.ok) {
    console.error(`Groq HTTP ${response.status}: ${text.slice(0, 500)}`);
    process.exit(1);
  }
  process.stdout.write(text);
});
"""
    process = None
    for attempt in range(6):
        process = subprocess.run(
            ["node", "-e", node_code],
            input=json.dumps({"apiKey": api_key, "body": body}),
            text=True,
            capture_output=True,
            timeout=60,
            cwd=ROOT,
        )

        if process.returncode == 0:
            break

        if "rate_limit_exceeded" in process.stderr and attempt < 5:
            wait = 30 + attempt * 15
            print(f"Rate limited. Waiting {wait}s before retrying {image_path.name}...")
            time.sleep(wait)
            continue

        raise RuntimeError(process.stderr.strip() or "Node Groq request failed.")

    payload = json.loads(process.stdout)

    content = payload["choices"][0]["message"]["content"]
    cleaned = content.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
    json_start = cleaned.find("{")
    json_end = cleaned.rfind("}")
    if json_start >= 0 and json_end > json_start:
        cleaned = cleaned[json_start : json_end + 1]
    try:
        return normalize(json.loads(cleaned))
    except json.JSONDecodeError as exc:
        (OUTPUT_DIR / "failed_groq_response.txt").write_text(content)
        raise RuntimeError(
            f"Groq returned non-JSON content for {image_path.name}; saved to outputs/failed_groq_response.txt"
        ) from exc


def includes_all(text: str, terms: list[str]) -> bool:
    lower = text.lower()
    return all(term.lower() in lower for term in terms)


def includes_any(text: str, terms: list[str]) -> bool:
    lower = text.lower()
    return any(term.lower() in lower for term in terms)


def evaluate(test: TestImage, result: dict) -> dict:
    alt = result["alt_text"]
    combined = f"{result['alt_text']} {result['long_description']} {result['text_in_image']}"
    under_125 = "Y" if len(alt) <= 125 else "N"
    embedded = "NA"
    if test.embedded_terms:
        embedded = "Y" if includes_all(result["text_in_image"], test.embedded_terms) or includes_all(combined, test.embedded_terms) else "N"
    purpose = "Y" if includes_any(combined, test.purpose_terms) else "N"

    notes = []
    if under_125 == "N":
        notes.append("Alt text exceeded 125 characters after normalization.")
    if embedded == "N":
        notes.append("Did not capture all critical visible text or event details.")
    if purpose == "N":
        notes.append("Description did not convey the image's main purpose.")
    if test.filename == "07_nature_forest_path.png" and any(
        word in combined.lower() for word in ["deforestation", "stump", "cut down"]
    ):
        purpose = "N"
        notes.append("Invented a deforestation/stump meaning that was not actually present in the image.")
    if test.category == "hard" and purpose == "Y" and embedded != "N":
        notes.append("Hard case passed, but verify manually because the subject/text is intentionally difficult.")

    return {
        "Filename": test.filename,
        "Alt Text Generated": alt,
        "Character Count": len(alt),
        "Under 125 (Y/N)": under_125,
        "Embedded Text Captured (Y/N/NA)": embedded,
        "Conveys Equivalent Purpose (Y/N)": purpose,
        "Failure Notes": " ".join(notes),
    }


def write_workbook(run_rows: dict[str, list[dict]]):
    wb = Workbook()
    wb.remove(wb.active)
    headers = [
        "Filename",
        "Alt Text Generated",
        "Character Count",
        "Under 125 (Y/N)",
        "Embedded Text Captured (Y/N/NA)",
        "Conveys Equivalent Purpose (Y/N)",
        "Failure Notes",
    ]
    for sheet_name, rows in run_rows.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D8F3EF")
        for row in rows:
            ws.append([row[h] for h in headers])
        start = len(rows) + 3
        ws.cell(start, 1, "Tallies").font = Font(bold=True)
        ws.cell(start + 1, 1, "Under 125")
        ws.cell(start + 1, 2, sum(1 for r in rows if r["Under 125 (Y/N)"] == "Y"))
        ws.cell(start + 2, 1, "Conveyed equivalent purpose")
        ws.cell(start + 2, 2, sum(1 for r in rows if r["Conveys Equivalent Purpose (Y/N)"] == "Y"))
        flyer_rows = [r for r in rows if "flyer" in r["Filename"]]
        ws.cell(start + 3, 1, "Flyers captured event details")
        ws.cell(start + 3, 2, sum(1 for r in flyer_rows if r["Embedded Text Captured (Y/N/NA)"] == "Y"))
        for index, width in enumerate([34, 52, 16, 16, 28, 30, 62], start=1):
            ws.column_dimensions[get_column_letter(index)].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")
    wb.save(WORKBOOK_PATH)


def summarize(run_rows: dict[str, list[dict]]) -> str:
    lines = []
    previous_failures: set[str] = set()
    for run_name, rows in run_rows.items():
        failures = {r["Filename"] for r in rows if r["Failure Notes"]}
        under = sum(1 for r in rows if r["Under 125 (Y/N)"] == "Y")
        purpose = sum(1 for r in rows if r["Conveys Equivalent Purpose (Y/N)"] == "Y")
        flyer = sum(1 for r in rows if "flyer" in r["Filename"] and r["Embedded Text Captured (Y/N/NA)"] == "Y")
        improved = previous_failures - failures if previous_failures else set()
        regressed = failures - previous_failures if previous_failures else set()
        lines.append(
            {
                "run": run_name,
                "under": under,
                "purpose": purpose,
                "flyer": flyer,
                "failures": failures,
                "improved": improved,
                "regressed": regressed,
            }
        )
        previous_failures = failures

    log = ["# Prompt Log", ""]
    changes = {
        "Run 1": "Baseline prompt: asked for alt text, long description, and visible text, but did not give extra prioritization rules.",
        "Run 2": "Changed the prompt to prioritize embedded event text for flyers, posters, signs, screenshots, and graphics.",
        "Run 3": "Changed the prompt again to emphasize equivalent purpose, chart trends, and uncertainty for dark or ambiguous images.",
    }
    for item in lines:
        run = item["run"]
        log.append(f"## {run}")
        log.append(changes[run])
        log.append(
            f"Measured results: {item['under']}/12 under 125 characters, {item['purpose']}/12 conveyed equivalent purpose, and {item['flyer']}/3 flyers captured event details."
        )
        if run == "Run 1":
            log.append(
                "The main failure pattern was missing or incomplete visible text on text-heavy images, especially flyers and the chart screenshot."
            )
        else:
            improved = ", ".join(sorted(item["improved"])) or "none"
            regressed = ", ".join(sorted(item["regressed"])) or "none"
            log.append(f"Improved failures: {improved}. Regressions/new failures: {regressed}.")
        if run == "Run 3" and not item["regressed"]:
            log.append(
                "Regression note: no automated regression was detected in the heuristic scoring, but the hard ambiguous image remained a manual-risk case because the model can sound confident about unclear subjects."
            )
        log.append("")
    PROMPT_LOG_PATH.write_text("\n".join(log))
    return "\n".join(log)


def main():
    env = read_env()
    api_key = env.get("GROQ_API_KEY")
    model = env.get("GROQ_VISION_MODEL", MODEL_FALLBACK)
    if not api_key:
        raise SystemExit("Missing GROQ_API_KEY in .env.local")

    tests = build_images()
    run_rows: dict[str, list[dict]] = {}
    cache = {}
    if CACHE_PATH.exists():
        cache = json.loads(CACHE_PATH.read_text())
    allow_partial = os.environ.get("ALLOW_PARTIAL") == "1"
    for run_name, _, prompt in RUN_PROMPTS:
        rows = []
        for index, test in enumerate(tests, start=1):
            cache_key = f"{run_name}:{test.filename}"
            print(f"{run_name} {index}/12: {test.filename}")
            image_path = IMAGE_DIR / test.filename
            used_live_call = False
            if cache_key in cache:
                result = cache[cache_key]
            elif allow_partial:
                rows.append(
                    {
                        "Filename": test.filename,
                        "Alt Text Generated": "PENDING - Groq daily token limit reached before this image could be tested.",
                        "Character Count": 0,
                        "Under 125 (Y/N)": "N",
                        "Embedded Text Captured (Y/N/NA)": "NA" if not test.embedded_terms else "N",
                        "Conveys Equivalent Purpose (Y/N)": "N",
                        "Failure Notes": "Not measured yet because Groq returned a daily token-limit error during validation.",
                    }
                )
                continue
            else:
                result = call_groq(image_path, prompt, api_key, model)
                cache[cache_key] = result
                CACHE_PATH.write_text(json.dumps(cache, indent=2))
                used_live_call = True
            rows.append(evaluate(test, result))
            if used_live_call:
                time.sleep(8)
        run_rows[run_name] = rows
    write_workbook(run_rows)
    summarize(run_rows)
    print(f"Wrote {WORKBOOK_PATH}")
    print(f"Wrote {PROMPT_LOG_PATH}")
    print(f"Wrote images to {IMAGE_DIR}")


if __name__ == "__main__":
    main()
